import openpyxl


class ExcelExporter:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None

    def create_workbook(self):
        self.workbook = openpyxl.Workbook()

    def write_data(self, title, data, sheet):
        ws = self.workbook.create_sheet(sheet)
        ws.append(title)
        for row in sorted(data, key=lambda x: x[1]):
            ws.append(row)

    def save_workbook(self):
        self.workbook.save(self.filename)

    # def get_sheet_name(self):
    #     wss = self.workbook.sheetnames
    #     return wss[0]

    def export_to_excel(self, title, data, sheet):
        # sheet = self.get_sheet_name()

        self.write_data(title, data, sheet)
        self.save_workbook()
